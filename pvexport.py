#!/usr/bin/env python3

# Proxmox VE export utility
# Version 1.0
# Copyright (C) 2026 Cyril Pawelko - Yumaiia
# https://www.yumaiia.fr/pvexport/
# GPLv3 license

## TODO - not implemented yet:
# Add support for API tokens
# Ceph output
# Networks details
# Hardwaire details via reports -> lscpu, dmidecode, 

import argparse
import getpass
from datetime import datetime
import os
import requests
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

SCRIPT_VERSION = "1.0"
PVE_DEFAULT_PORT = 8006
COLOR_NAVY = "1B395C"
COLOR_CYAN = "64B9C9"

def build_base_url(server):
    """Accepts either a full URL (https://server:port) or a bare server/IP,
    in which case https:// and the default Proxmox port are assumed."""
    if server.startswith("http://") or server.startswith("https://"):
        return server.rstrip("/")
    return f"https://{server}:{PVE_DEFAULT_PORT}"

def parse_args():
    parser = argparse.ArgumentParser(description="Proxmox VE export utility")
    parser.add_argument("-s", "--server", required=True,
                         help="Proxmox VE cluster URL or IP/hostname "
                              "(e.g. https://pve01.example.local:8006 or 192.168.20.170)")
    parser.add_argument("-u", "--user", required=True,
                         help="Proxmox VE username (e.g. root@pve). PVE.Auditor permission are required")
    parser.add_argument("-p", "--password",
                         help="Proxmox VE password (omit to be prompted)")
    parser.add_argument("-i", "--insecure", action="store_true",
                         help="Do not verify the server's TLS certificate (insecure)")
    parser.add_argument("-o", "--output",
                         help="Output Excel filename "
                              "(default: pvexport-<date>-<time>.xlsx)")
    parser.add_argument("-n", "--no-open", action="store_true",
                         help="Do not automatically open the generated file")
    parser.add_argument("-v", "--verbose", action="store_true",
                         help="Print detailed debug output (full data dumps)")
    return parser.parse_args()

args = parse_args()
user = args.user

if user.lower() == "root@pam":
    print(
        "Warning: you are using 'root@pam', a full administrative account. "
        "It is recommended to use a dedicated read-only account instead "
        "(e.g. with the PVE.Auditor role)."
    )

password = args.password or getpass.getpass(f"Password for {user}: ")
baseurl = build_base_url(args.server)
apiurl = baseurl + "/api2/json"
insecure = args.insecure

if insecure:
    requests.packages.urllib3.disable_warnings()

xl_filename = args.output or f"pvexport-{datetime.now():%Y%m%d-%H%M%S}.xlsx"

def get_ticket(user, password):
    url = apiurl + "/access/ticket"
    data = {"username": user, "password": password}
    try:
        response = requests.post(url, data=data, verify=not insecure)
        response.raise_for_status()
        return response.json()["data"]["ticket"], response.json()["data"]["CSRFPreventionToken"]
    except requests.exceptions.SSLError as e:
        raise SystemExit(
            f"Error occurred while fetching ticket: {e}\n"
            "Hint: if the server uses a self-signed or untrusted certificate, "
            "retry with the -i/--insecure option."
        ) from e
    except requests.RequestException as e:
        raise SystemExit(f"Error occurred while fetching ticket: {e}") from e

def get_data(data_url, ticket, csrf_token):
    url = apiurl + data_url
    headers = {
        "cookie": f"PVEAuthCookie={ticket}",
        "CSRFPreventionToken": csrf_token }   
    try:
        response = requests.get(url, headers=headers, verify=not insecure)
        response.raise_for_status()
        if "data" in response.json():
            return response.json()["data"]
        else:
            print(f"No data found in response for {data_url}")
            return None
    except requests.exceptions.SSLError as e:
        print(
            f"Error occurred while fetching data: {e}\n"
            "Hint: if the server uses a self-signed or untrusted certificate, "
            "retry with the -i/--insecure option."
        )
        return None
    except requests.RequestException as e:
        print(f"Error occurred while fetching data: {e}")
        return None

def get_value(data, path, default=""):
    for key in path.split("."):
        if not isinstance(data, dict):
            return default
        data = data.get(key)
        if data is None:
            return default
    return data

# Style definitions
header_style = openpyxl.styles.NamedStyle(name="header_style")
header_style.font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
header_style.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color=COLOR_NAVY)
header_style.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
header_style.border = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style="medium", color="000000"),
    left=openpyxl.styles.Side(border_style="thin"),
    right=openpyxl.styles.Side(border_style="thin"),
    top=openpyxl.styles.Side(border_style="thin")
)

data_style = openpyxl.styles.NamedStyle(name="data_style")
data_style.font = openpyxl.styles.Font(size=11, color="000000")
data_style.border = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style="thin", color="000000"),
    left=openpyxl.styles.Side(border_style="thin"),
    right=openpyxl.styles.Side(border_style="thin"),
    top=openpyxl.styles.Side(border_style="thin")
)

def write_sheet(wb, name, data, fields):
    ws = wb.create_sheet(title=name)
    col_index = 1
    for header in fields.values():
        cell = ws.cell(1, col_index)
        cell.value = header["label"]
        cell.style = header_style
        col_index += 1
    row_index = 2
    for item in data:
        col_index = 1
        for meta in fields.values():
            #value = item.get(meta["source"], "")
            #value = reduce(getitem, meta["source"].split("."), item)
            value = get_value(item, meta["source"], "")
            if "transform" in meta and callable(meta["transform"]):
                value = meta["transform"](value)
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.style = data_style
            if meta.get("cell_format"):
                cell.number_format = meta["cell_format"]
            col_index += 1
        row_index += 1
    # Auto-size columns
    for col_index in ws.columns:
        max_length = 0
        column = col_index[0].column_letter
        for cell in col_index:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column].width = adjusted_width

# Transformations definitions
b_to_mib = lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None
b_to_gib = lambda x: round(x / (1024 * 1024 * 1024), 2) if (x is not None) and (x != '') else None
yes_no = lambda x: "Yes" if x else "No"
days = lambda x: x / 86400 if (x is not None) and (x != '') else None
to_int = lambda x: int(x) if (x is not None) and (x != '') else None
round0 = lambda x: round(x, 0) if (x is not None) and (x != '') else None
round2 = lambda x: round(x, 2) if (x is not None) and (x != '') else None

ticket,csrf_token = get_ticket(user, password)
#print("ticket = ", ticket)
#print("CSRF Token = ", csrf_token)
print("Connected to Proxmox VE API" )

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Info"

# Get cluster status
cluster_fields = {
    "name": {
        "source" : "name",
        "label" : "Name",
    },
    "version": {
        "source" : "version",
        "label" : "Corosync config version",
    },
    "nodes": {
        "source" : "nodes",
        "label" : "Nodes",
    },
    "quorate": {
        "source" : "quorate",
        "label" : "Quorate",
        "transform": yes_no,
    },
}

cluster_status = [next((item for item in get_data("/cluster/status", ticket, csrf_token) if item.get("type") == "cluster"), None)]
write_sheet(wb, "Cluster", cluster_status, cluster_fields)

# Populate Info sheet
YUMAIIA_LINK = "https://www.yumaiia.fr/pvexport/"

# Brand banner (styled after the Yumaiia logo: navy background, white wordmark,
# teal accent bar) -- text only, no image asset involved.
navy_fill = openpyxl.styles.PatternFill(fill_type="solid", start_color=COLOR_NAVY)
teal_fill = openpyxl.styles.PatternFill(fill_type="solid", start_color=COLOR_CYAN)
banner_center = openpyxl.styles.Alignment(horizontal="center", vertical="center")

ws.merge_cells("A1:B1")
title_cell = ws["A1"]
title_cell.value = CellRichText(
    TextBlock(InlineFont(b=True, sz="22", color="FFFFFF", rFont="Segoe UI"), "pve"),
    TextBlock(InlineFont(b=True, sz="22", color=COLOR_CYAN, rFont="Segoe UI"), "X"),
    TextBlock(InlineFont(b=True, sz="22", color="FFFFFF", rFont="Segoe UI"), "port"),
)
title_cell.font = openpyxl.styles.Font(bold=True, size=22, color="FFFFFF", name="Segoe UI")
title_cell.fill = navy_fill
title_cell.alignment = banner_center
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:B2")
subtitle_cell = ws["A2"]
subtitle_cell.value = CellRichText(
    TextBlock(InlineFont(b=True, sz="11", color="FFFFFF", rFont="Segoe UI"), "by Yuma"),
    TextBlock(InlineFont(b=True, sz="11", color=COLOR_CYAN, rFont="Segoe UI"), "ii"),
    TextBlock(InlineFont(b=True, sz="11", color="FFFFFF", rFont="Segoe UI"), "a"),
)
subtitle_cell.font = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF", name="Segoe UI")
subtitle_cell.fill = navy_fill
subtitle_cell.alignment = banner_center
ws.row_dimensions[2].height = 18

ws.merge_cells("A3:B3")
tagline_cell = ws["A3"]
tagline_cell.value = YUMAIIA_LINK
tagline_cell.hyperlink = YUMAIIA_LINK
tagline_cell.font = openpyxl.styles.Font(size=11, color=COLOR_CYAN, underline="single", name="Segoe UI")
tagline_cell.fill = navy_fill
tagline_cell.alignment = banner_center
ws.row_dimensions[3].height = 18

ws.merge_cells("A4:B4")
ws["A4"].fill = teal_fill
ws.row_dimensions[4].height = 5

info_rows = [
    ("pveXport version", SCRIPT_VERSION),
    ("Server URL", baseurl),
    ("Generated on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
]

label_font = openpyxl.styles.Font(bold=True)

for offset, (label, value) in enumerate(info_rows, start=1):
    row_index = 4 + offset
    label_cell = ws.cell(row=row_index, column=1, value=label)
    label_cell.font = label_font
    ws.cell(row=row_index, column=2, value=value)

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 50

# Get nodes information
# Sample data:
# {"data":{"ksm":{"shared":150302720},"kversion":"Linux 6.17.4-1-pve #1 SMP PREEMPT_DYNAMIC PMX 6.17.4-1 (2025-12-03T15:42Z)","idle":0,"memory":{"free":1277120512,"available":3868946432,"total":16540520448,"used":12671574016},"loadavg":["1.74","1.86","1.50"],"wait":0.00537015726889145,"boot-info":{"secureboot":0,"mode":"efi"},"swap":{"used":6283759616,"total":8589930496,"free":2306170880},"pveversion":"pve-manager/9.1.2/9d436f37a0ac4172","cpu":0.404739747502543,"rootfs":{"avail":71318425600,"used":24372563968,"total":100861726720,"free":76489162752},"cpuinfo":{"sockets":1,"mhz":"2895.985","hvm":"1","user_hz":100,"flags":"fpu vme de pse tsc msr pae mce cx8 apic sep mtrr pge mca cmov pat pse36 clflush dts acpi mmx fxsr sse sse2 ss ht tm pbe syscall nx pdpe1gb rdtscp lm constant_tsc art arch_perfmon pebs bts rep_good nopl xtopology nonstop_tsc cpuid aperfmperf tsc_known_freq pni pclmulqdq dtes64 monitor ds_cpl vmx est tm2 ssse3 sdbg fma cx16 xtpr pdcm pcid sse4_1 sse4_2 x2apic movbe popcnt tsc_deadline_timer aes xsave avx f16c rdrand lahf_lm abm 3dnowprefetch cpuid_fault epb cat_l2 cdp_l2 ssbd ibrs ibpb stibp ibrs_enhanced tpr_shadow flexpriority ept vpid ept_ad fsgsbase tsc_adjust bmi1 avx2 smep bmi2 erms invpcid rdt_a rdseed adx smap clflushopt clwb intel_pt sha_ni xsaveopt xsavec xgetbv1 xsaves split_lock_detect user_shstk avx_vnni dtherm ida arat pln pts hwp hwp_notify hwp_act_window hwp_epp hwp_pkg_req vnmi umip pku ospke waitpkg gfni vaes vpclmulqdq rdpid movdiri movdir64b fsrm md_clear serialize arch_lbr ibt flush_l1d arch_capabilities","model":"Intel(R) N100","cores":4,"cpus":4},"uptime":1480737,"current-kernel":{"machine":"x86_64","release":"6.17.4-1-pve","sysname":"Linux","version":"#1 SMP PREEMPT_DYNAMIC PMX 6.17.4-1 (2025-12-03T15:42Z)"}}}

nodes_fields = {
    "nodeid": {
        "source": "nodeid",
        "label": "ID",
        "transform": to_int,
    },
    "name": {
        "source": "name",
        "label": "Name"
    },
    "status": {
        "source": "status",
        "label": "Status"
    },
     "pve_version": {
        "source": "pveversion",
        "label": "PVE Version"
    },
    "quorum_votes": {
        "source": "quorum_votes",
        "label": "Quorum Votes",
        "transform": to_int,
    },
    "hastate": {
        "source": "hastate",
        "label": "HA state"
    },
    "kernel_machine": {
        "source": "current-kernel.machine",
        "label": "Kernel machine"
    },
    "kernel_release": {
        "source": "current-kernel.release",
        "label": "Kernel release"
    },
    "boot_mode": {
        "source": "boot-info.mode",
        "label": "Boot mode"
    },
    "secure_boot": {
        "source": "boot-info.secureboot",
        "label": "Secure boot",
        "transform": yes_no
    },
    "ring0_addr": {
        "source": "ring0_addr",
        "label": "Ring0 address"
    },
   "uptime": {
        "source": "uptime",
        "label": "Uptime (s)"
    },
    "uptime_days": {
        "source": "uptime",
        "label": "Uptime (d)",
        "transform": days,
        "cell_format": r"dd\d hh:mm:ss"
    },
    "cpu_model": {
        "source": "cpuinfo.model",
        "label": "CPU Model"
    },
    "cpu_mhz": {
        "source": "cpuinfo.mhz",
        "label": "CPU MHz",
    },
    "cpu_cores": {
        "source": "cpuinfo.cores",
        "label": "Cores per CPU"
    },
    "cpu_sockets": {
        "source": "cpuinfo.sockets",
        "label": "CPU Sockets"
    },
    "cpu_cpus": {
        "source": "cpuinfo.cpus",
        "label": "CPUs cores"
    },
    "loadavg_1": {
        "source": "loadavg",
        "label": "Load avg. (1m)",
        "transform": lambda x: x[0] if isinstance(x, list) and len(x) > 0 else None,
    },
    "loadavg_5": {
        "source": "loadavg",
        "label": "Load avg. (5m)",
        "transform": lambda x: x[1] if isinstance(x, list) and len(x) > 1 else None,
    },
    "loadavg_15": {
        "source": "loadavg",
        "label": "Load avg. (15m)",
        "transform": lambda x: x[2] if isinstance(x, list) and len(x) > 2 else None,
    },
    "memory_total": {
        "source": "memory.total",
        "label": "Memory Total (B)"
    },
    "memory_total_mib": {
        "source": "memory.total",
        "label": "Memory Total (MiB)",
        "transform": b_to_mib, #lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None,
        "cell_format": "0.00"
    },
    "memory_used": {
        "source": "memory.used",
        "label": "Memory Used (B)"
    },
    "memory_used_mib": {
        "source": "memory.used",
        "label": "Memory Used (MiB)",
        "transform": b_to_mib, #lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None,
        "cell_format": "0.00"
    },
    "memory_free": {
        "source": "memory.free",
        "label": "Memory Free (B)"
    },
    "memory_free_mib": {
        "source": "memory.free",
        "label": "Memory Free (MiB)",
        "transform": b_to_mib, #lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None,
        "cell_format": "0.00"
    },
    "memory_available": {
        "source": "memory.available",
        "label": "Memory Available (B)"
    },
    "memory_available_mib": {
        "source": "memory.available",
        "label": "Memory Available (MiB)",
        "transform": b_to_mib, #lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None,
        "cell_format": "0.00"
    },
    "ksm_shared": {
        "source": "ksm.shared",
        "label": "KSM Shared (B)"
    },
     "ksm_shared_mib": {
        "source": "ksm.shared",
        "label": "KSM Shared (MiB)",
        "transform": b_to_mib, #lambda x: round(x / (1024 * 1024), 2) if (x is not None) and (x != '') else None,
        "cell_format": "0.00"
    },
}

nodes_resources = get_data("/cluster/resources?type=node", ticket, csrf_token)
nodes_list = get_data("/nodes", ticket, csrf_token)
nodes_config =  get_data("/cluster/config/nodes", ticket, csrf_token)

nodes = []
for node in nodes_resources:
    node_status = get_data(f"/nodes/{node['node']}/status", ticket, csrf_token)
    config = next((item for item in nodes_config if item.get("name") == node["node"]), {})
    #print("config:", config)
    nodes.append(node | node_status | config)

if args.verbose:
    print(nodes)
write_sheet(wb, "Nodes", nodes, nodes_fields)

# Get VMs (lxc + qemu) resources
vm_fields = {
    "vmid": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "vmid",
        "label" : "ID",
    },
    "type": {
        "target": ["vm"],
        "source" : "type",
        "label" : "Type",
    },
    "arch": {
        "target": ["lxc"],
        "source" : "arch", 
        "label" : "Arch",
    },
    "name": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "name",
        "label" : "Name",
    },
    "status": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "status",
        "label" : "Status",
    },
    "maxcpu": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "maxcpu",
        "label" : "Max CPU",
    },
    "maxmem": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "maxmem",
        "label" : "Max mem. (B)",
    },
    "maxmem_mib": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "maxmem",
        "label" : "Max mem. (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024)) if (x is not None) and (x != '') else None,
        "cell_format" : "0",
    },
    "maxdisk": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "maxdisk",
        "label" : "Max disk (B)",
    },
    "maxdisk_mib": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "maxdisk",
        "label" : "Max disk (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024), 0) if (x is not None) and (x != '') else None,
    },
    "swap": {
        "target": ["lxc"],
        "source" : "swap", 
        "label" : "Swap",
    },
    "swap_mib": {
        "target": ["lxc"],
        "source" : "swap",
        "label" : "Swap (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024), 0) if (x is not None) and (x != '') else None,
    },
    "cpu": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "cpu",
        "label" : "CPU usage",
        "transform" : round2, #lambda x: round(x, 2) if (x is not None) and (x != '') else None,
        "cell_format" : "0%",
    },
    "mem": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "mem",
        "label" : "Mem. used (B)",
    },
    "mem_mib": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "mem",
        "label" : "Mem. used (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024), 0) if (x is not None) and (x != '') else None,
    },
    "memhost": {
        "target": ["vm", "qemu"],
        "source" : "memhost",
        "label" : "Host mem. used (B)",
    },
    "memhost_mib": {
        "target": ["vm", "qemu"],
        "source" : "memhost",
        "label" : "Host mem. used (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024), 0) if (x is not None) and (x != '') else None,
    },
   "rootfs": {
        "target": ["lxc"],
        "source" : "rootfs",
        "label" : "Rootfs",
    },
    "disk": {
        "target": ["lxc"],
        "source" : "disk",
        "label" : "Disk used (B)",
    },
    "disk_mib": {
        "target": ["lxc"],
        "source" : "disk",
        "label" : "Disk used (MiB)",
        "transform" : b_to_mib, #lambda x: round(x / (1024 * 1024), 0) if (x is not None) and (x != '') else None,
    },    
    "uptime": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "uptime",
        "label" : "Uptime (s)",
    },
    "uptime_days": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "uptime",
        "label" : "Uptime (d)",
        "transform": days,
        "cell_format" : r"dd\d hh:mm:ss", #mm\mjj\dhh:mm
    },
    "pool": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "pool",
        "label" : "Pool",
    },
    "node": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "node",
        "label" : "Node",
    },
    "hastate": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "hastate",
        "label" : "HA state",
    },
    "ha" : {
        "target": ["lxc", "qemu"],
        "source" : "ha",
        "label" : "HA",
        "transform" : lambda x: ", ".join(f"{k}={v}" for k, v in x.items()) if isinstance(x, dict) else str(x),
    },
    "tags": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "tags",
        "label" : "Tags",
    },
    "template": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "template",
        "label" : "Is template",
        "transform" : yes_no,
    },
    "features": {
        "target": ["lxc"],
        "source" : "features",
        "label" : "Features",
    },
    "onboot": {
        "target": ["vm", "lxc", "qemu"],
        "source" : "onboot",
        "label" : "On boot",
        "transform" : yes_no,
    },
    "unprivileged": {
        "target": ["lxc"],
        "source" : "unprivileged",
        "label" : "Unprivileged",
        "transform" : yes_no,
    },
    "net0": {
        "target": ["lxc", "qemu"],
        "source" : "net0",
        "label" : "Net0",
    },
    "net1": {
        "target": ["lxc", "qemu"],
        "source" : "net1",
        "label" : "Net1",
    },
    "net2": {
        "target": ["lxc", "qemu"],
        "source" : "net2",
        "label" : "Net2",
    },
    "net3": {
        "target": ["lxc", "qemu"],
        "source" : "net3",
        "label" : "Net3",
    },   
    "bios": {
        "target": ["qemu"],
        "source" : "bios",
        "label" : "BIOS",
    },
    "ostype": {
        "target": ["qemu"],
        "source" : "ostype",
        "label" : "OS type",
    },
    "scsihw": {
        "target": ["qemu"],
        "source" : "scsihw",
        "label" : "SCSI HW",
    },
    "agent": {
        "target": ["qemu"], 
        "source" : "agent",
        "label" : "QEMU agent",
        "transform" : yes_no,
    },
    "balloon": {
        "target": ["qemu"],
        "source" : "balloon",
        "label" : "Balloon",
    },
    "running-machine": {
        "target": ["qemu"],
        "source" : "running-machine",
        "label" : "Running Machine",
    },
    "running-qemu": {
        "target": ["qemu"],
        "source" : "running-qemu",
        "label" : "Running qemu",
    },
    "freemem": {
        "target": ["qemu"],
        "source" : "freemem",
        "label" : "Free mem. (B)",
    },
}

resources = get_data("/cluster/resources?type=vm", ticket, csrf_token)
resources_fields = {k: v for k, v in vm_fields.items() if "vm" in v.get("target", [])}
write_sheet(wb, "Resources", resources, resources_fields)

## LXC and QEMU details
lxc=[]
qemu=[]
for resource in resources:
    if resource["type"] == "lxc":
        config=get_data(f"/nodes/{resource['node']}/lxc/{resource['vmid']}/config", ticket, csrf_token)
        current_status=get_data(f"/nodes/{resource['node']}/lxc/{resource['vmid']}/status/current", ticket, csrf_token)
        lxc.append(resource | (config or {}) | (current_status or {}))

    elif resource["type"] == "qemu":
        config=get_data(f"/nodes/{resource['node']}/qemu/{resource['vmid']}/config", ticket, csrf_token)
        current_status=get_data(f"/nodes/{resource['node']}/qemu/{resource['vmid']}/status/current", ticket, csrf_token)
        qemu.append(resource | (config or {}) | (current_status or {}))

lxc_fields = {k: v for k, v in vm_fields.items() if "lxc" in v.get("target", [])}
write_sheet(wb, "LXC", lxc, lxc_fields)

if args.verbose:
    print("qemu details:", qemu)
qemu_fields = {k: v for k, v in vm_fields.items() if "qemu" in v.get("target", [])}
write_sheet(wb, "QEMU", qemu, qemu_fields)

# Disks
disk_fields = {
        "node": {
            "source": "node",
            "label": "Node",
        },
        "dev_path": {
            "source": "devpath",
            "label": "Device path",
        },
        "by_id_link": {
            "source": "by_id_link",
            "label": "By ID link",
        },
        "vendor": {
            "source": "vendor",
            "label": "Vendor",
        },
        "model": {
            "source": "model",
            "label": "Model",
        },
        "type": {
            "source": "type",
            "label": "Type",
        },
        "serial": {
            "source": "serial",
            "label": "Serial",
        },
        "used": {
            "source": "used",
            "label": "Used (B)",
        },
        "health": {
            "source": "health",
            "label": "Health",
        },
        "wearout": {
            "source": "wearout",
            "label": "Wearout",
        },
        "wwn": {
            "source": "wwn",
            "label": "WWN",
        },
        "size": {
            "source": "size",
            "label": "Size (B)",
        },
        "size_gib": {
            "source": "size",
            "label": "Size (GiB)",
            "transform": b_to_gib, #lambda x: round(x / (1024 * 1024 * 1024), 2) if (x is not None) and (x != '') else None,
            "cell_format": "0.00"
        },
        "osdid": {
            "source": "osdid",
            "label": "OSD ID",
        },
}

disks = []
for node in nodes_resources:
    node_disks = get_data(f"/nodes/{node['node']}/disks/list", ticket, csrf_token)
    #    node_status = get_data(f"/nodes/{node['node']}/status", ticket, csrf_token)
    for disk in node_disks:
        disk["node"] = node["node"]
        disks.append(disk)

if args.verbose:
    print("disks:", disks)
write_sheet(wb, "Disks", disks, disk_fields)

wb.active = wb["Info"]
wb.save(xl_filename)

if not args.no_open and os.name == 'nt':
    os.startfile(xl_filename)